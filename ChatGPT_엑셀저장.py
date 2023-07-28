import random
from faker import Faker
from openpyxl import Workbook
from openpyxl.utils import get_column_letter

# Initialize faker
fake = Faker()

# Create a workbook and select the active worksheet
wb = Workbook()
ws = wb.active

# Define the headers
headers = ['ProductID', 'ProductName', 'Price', 'SaleDate']

# Write the headers to the worksheet
for col_num, header in enumerate(headers, 1):
    col_letter = get_column_letter(col_num)
    ws['{}1'.format(col_letter)] = header
    ws.column_dimensions[col_letter].width = 15

# Generate data
for row_num in range(2, 1002):  # 1000 records
    product_id = row_num - 1  # unique product id
    product_name = fake.word(ext_word_list=None)  # random product name
    price = round(random.uniform(50, 5000), 2)  # random price between 50 and 5000
    sale_date = fake.date_between(start_date='-3y', end_date='today')  # random sale date in last three years

    ws.append([product_id, product_name, price, sale_date])

# Save the workbook
wb.save("sales_data.xlsx")
